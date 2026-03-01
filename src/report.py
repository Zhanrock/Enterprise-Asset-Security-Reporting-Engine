"""
report.py - Generate multi-sheet Excel dashboards with Pivot-style summaries,
            VLOOKUP equivalent tables, and compliance/patch views.
"""

import os
import sqlite3
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from database import get_connection, DB_PATH

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs")


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def _style_header_row(ws, row: int, fill_color: str = "2E75B6") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4


def _thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------

def _load_assets(conn) -> pd.DataFrame:
    return pd.read_sql_query("SELECT * FROM Assets", conn)


def _load_vulnerabilities(conn) -> pd.DataFrame:
    return pd.read_sql_query(
        """
        SELECT v.*, a.name AS asset_name, a.category, a.owner
        FROM Vulnerabilities v
        LEFT JOIN Assets a ON a.asset_id = v.asset_id
        """,
        conn,
    )


def _load_compliance(conn) -> pd.DataFrame:
    return pd.read_sql_query(
        """
        SELECT c.*, a.name AS asset_name, a.category
        FROM Compliance_Status c
        LEFT JOIN Assets a ON a.asset_id = c.asset_id
        """,
        conn,
    )


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _write_assets_sheet(wb, df_assets: pd.DataFrame) -> None:
    ws = wb.create_sheet("Assets Inventory")
    headers = list(df_assets.columns)
    ws.append(headers)
    _style_header_row(ws, 1)
    for _, row in df_assets.iterrows():
        ws.append(list(row))
    _auto_width(ws)


def _write_vuln_sheet(wb, df_vuln: pd.DataFrame) -> None:
    ws = wb.create_sheet("Vulnerabilities")
    headers = list(df_vuln.columns)
    ws.append(headers)
    _style_header_row(ws, 1, "C00000")

    # Colour-code severity
    severity_colours = {
        "Critical": "FF0000",
        "High":     "FF6600",
        "Medium":   "FFCC00",
        "Low":      "92D050",
    }
    for _, row in df_vuln.iterrows():
        ws.append(list(row))
        row_idx = ws.max_row
        sev = str(row.get("severity", ""))
        if sev in severity_colours:
            fill = PatternFill("solid", fgColor=severity_colours[sev])
            for cell in ws[row_idx]:
                cell.fill = fill
    _auto_width(ws)


def _write_compliance_sheet(wb, df_comp: pd.DataFrame) -> None:
    ws = wb.create_sheet("Compliance Status")
    headers = list(df_comp.columns)
    ws.append(headers)
    _style_header_row(ws, 1, "375623")
    for _, row in df_comp.iterrows():
        ws.append(list(row))
        row_idx = ws.max_row
        if row.get("compliant") == 1:
            fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            fill = PatternFill("solid", fgColor="FFC7CE")
        for cell in ws[row_idx]:
            cell.fill = fill
    _auto_width(ws)


def _write_pivot_summary(wb, df_assets: pd.DataFrame,
                         df_vuln: pd.DataFrame,
                         df_comp: pd.DataFrame) -> None:
    ws = wb.create_sheet("KPI Dashboard")

    # --- Asset summary pivot ---
    ws["A1"] = "ASSET SUMMARY BY CATEGORY"
    ws["A1"].font = Font(bold=True, size=14)

    pivot_assets = (
        df_assets.groupby("category")
        .agg(total=("asset_id", "count"),
             active=("status", lambda x: (x == "Active").sum()))
        .reset_index()
    )
    pivot_assets.columns = ["Category", "Total Assets", "Active"]
    ws.append([])
    headers = list(pivot_assets.columns)
    ws.append(headers)
    _style_header_row(ws, ws.max_row)
    for _, r in pivot_assets.iterrows():
        ws.append(list(r))

    asset_pivot_end = ws.max_row

    # --- Vulnerability summary pivot ---
    ws.append([])
    ws.append(["VULNERABILITY SUMMARY BY SEVERITY"])
    ws[ws.max_row][0].font = Font(bold=True, size=14)

    pivot_vuln = (
        df_vuln.groupby("severity")
        .agg(total=("vuln_id", "count"),
             open_=("status", lambda x: (x == "Open").sum()))
        .reset_index()
    )
    pivot_vuln.columns = ["Severity", "Total", "Open"]
    ws.append(list(pivot_vuln.columns))
    _style_header_row(ws, ws.max_row, "C00000")
    for _, r in pivot_vuln.iterrows():
        ws.append(list(r))

    # --- Compliance KPI ---
    ws.append([])
    ws.append(["COMPLIANCE KPI"])
    ws[ws.max_row][0].font = Font(bold=True, size=14)
    total_controls = len(df_comp)
    compliant = (df_comp["compliant"] == 1).sum()
    pct = round(compliant / total_controls * 100, 1) if total_controls else 0
    ws.append(["Total Controls Checked", total_controls])
    ws.append(["Compliant", compliant])
    ws.append(["Compliance %", f"{pct}%"])

    _auto_width(ws)

    # Add bar chart for vulnerability severity
    if len(pivot_vuln) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Open Vulnerabilities by Severity"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Severity"

        # Find the vuln pivot start row
        # Simple approach: search for "Severity" header
        sev_row = None
        for row in ws.iter_rows():
            if row[0].value == "Severity":
                sev_row = row[0].row
                break
        if sev_row:
            data_ref = Reference(ws, min_col=3, min_row=sev_row,
                                 max_row=sev_row + len(pivot_vuln))
            cats_ref = Reference(ws, min_col=1, min_row=sev_row + 1,
                                 max_row=sev_row + len(pivot_vuln))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            ws.add_chart(chart, "E2")


def _write_patch_compliance_vlookup(wb, df_assets: pd.DataFrame,
                                    df_vuln: pd.DataFrame) -> None:
    """
    Produce an 'INDEX/MATCH-style' lookup table that shows each asset,
    its open vulnerability count, and a patch compliance flag.
    (Pure pandas equivalent of Excel VLOOKUP/INDEX MATCH.)
    """
    ws = wb.create_sheet("Patch Compliance")

    open_counts = (
        df_vuln[df_vuln["status"] == "Open"]
        .groupby("asset_id")["vuln_id"]
        .count()
        .rename("open_vulns")
        .reset_index()
    )

    merged = df_assets.merge(open_counts, on="asset_id", how="left")
    merged["open_vulns"] = merged["open_vulns"].fillna(0).astype(int)
    merged["patch_compliant"] = merged["open_vulns"].apply(
        lambda x: "Yes" if x == 0 else "No"
    )

    cols = ["asset_id", "name", "category", "owner",
            "warranty_expiry", "status", "open_vulns", "patch_compliant"]
    ws.append(cols)
    _style_header_row(ws, 1)
    for _, row in merged[cols].iterrows():
        ws.append(list(row))
        row_idx = ws.max_row
        if row["patch_compliant"] == "Yes":
            fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            fill = PatternFill("solid", fgColor="FFC7CE")
        ws.cell(row=row_idx, column=8).fill = fill
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_report(db_path: str = DB_PATH, output_dir: str = OUTPUT_DIR) -> str:
    os.makedirs(output_dir, exist_ok=True)
    conn = get_connection(db_path)

    df_assets = _load_assets(conn)
    df_vuln   = _load_vulnerabilities(conn)
    df_comp   = _load_compliance(conn)
    conn.close()

    if df_assets.empty:
        print("[Report] No asset data found. Run ingest.py first.")
        return ""

    # Create workbook
    import openpyxl
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _write_pivot_summary(wb, df_assets, df_vuln, df_comp)
    _write_patch_compliance_vlookup(wb, df_assets, df_vuln)
    _write_assets_sheet(wb, df_assets)
    _write_vuln_sheet(wb, df_vuln)
    _write_compliance_sheet(wb, df_comp)

    filename = f"asset_security_report_{date.today().isoformat()}.xlsx"
    out_path = os.path.join(output_dir, filename)
    wb.save(out_path)
    print(f"[Report] Saved to {out_path}")
    return out_path


if __name__ == "__main__":
    generate_report()
